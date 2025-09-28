import os
import json
import logging
from datetime import datetime, timezone
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

# Constants for test classification
TEST_AREAS = {
    'auth': 'Authentication',
    'filter': 'Data Filtering', 
    'booking': 'Booking Management',
    'crud': 'Booking Management',
    'concurrent': 'Concurrent Operations',
    'health': 'Health Check',
    'ping': 'Health Check'
}

SEVERITY_MARKERS = {
    'critical': 'Critical',
    'smoke': 'High'
}

DEFAULT_TEST_AREA = 'General'
DEFAULT_SEVERITY = 'Medium'


class BugReporter:
    def __init__(self):
        self.bugs: List[Dict[str, Any]] = []
    
    def add_bug(self, expected: str, actual: str, severity: str = "Medium", 
                test_name: str = "", bug_type: str = "Auto-detected"):
        """Add a bug report"""
        bug = {
            "id": f"BUG-{len(self.bugs) + 1:03d}",
            "test_name": test_name,
            "test_area": self.get_test_area_from_name(test_name),
            "expected": expected,
            "actual": actual,
            "severity": severity,
            "bug_type": bug_type,
            "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            "environment": os.getenv('TEST_ENV', 'prod'),
            "status": "Open"
        }
        self.bugs.append(bug)
        return bug["id"]
    
    def get_test_area_from_name(self, test_name: str) -> str:
        """Extract test area from test name using constants"""
        test_name_lower = test_name.lower()
        
        for keyword, area in TEST_AREAS.items():
            if keyword in test_name_lower:
                return area
        
        return DEFAULT_TEST_AREA
    
    def get_severity_from_name(self, test_name: str) -> str:
        """Extract severity from test name patterns using constants"""
        test_name_lower = test_name.lower()
        
        for marker, severity in SEVERITY_MARKERS.items():
            if marker in test_name_lower:
                return severity
        
        return DEFAULT_SEVERITY
    
    def add_auto_detected_bug(self, test_name: str, failure_message: str, 
                            test_item=None, expected_behavior: str = None):
        """Add an automatically detected bug from test failure"""
        severity = self.get_severity_from_name(test_name)
        
        # Extract expected behavior from test docstring or use default
        if not expected_behavior:
            expected_behavior = self.get_test_description(test_item) or "Test should pass"
        
        # Clean up failure message
        if len(failure_message) > 500:
            failure_message = failure_message[:500] + "..."
        failure_message = failure_message.replace('\x00', '').replace('\n', ' ').replace('\r', ' ')
        
        return self.add_bug(
            expected=expected_behavior,
            actual=failure_message,
            severity=severity,
            test_name=test_name,
            bug_type="Auto-detected"
        )
    
    def get_test_description(self, test_item):
        """Extract test description from docstring or function name"""
        if not test_item:
            return None
            
        # Try to get docstring from the test function
        if hasattr(test_item, 'function') and test_item.function.__doc__:
            docstring = test_item.function.__doc__.strip()
            # Return first line of docstring (usually the description)
            return docstring.split('\n')[0].strip()
        
        # Fallback: convert test name to readable description
        test_name = test_item.name if hasattr(test_item, 'name') else str(test_item)
        # Convert test_filter_by_name -> "Filter by name should work"
        readable_name = test_name.replace('test_', '').replace('_', ' ').title()
        return f"{readable_name} should work correctly"
    
    def generate_excel_report(self, filename: str = None):
        """Generate enhanced Excel bug report"""
        if not filename:
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"reports/enhanced_bug_report_{timestamp}.xlsx"
        
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bug Report"
        
        # Headers
        headers = ["Bug ID", "Test Name", "Test Area", "Severity", "Bug Type", 
                  "Expected", "Actual", "Environment", "Status", "Timestamp"]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add bug data
        for row, bug in enumerate(self.bugs, 2):
            ws.cell(row=row, column=1, value=bug["id"])
            ws.cell(row=row, column=2, value=bug["test_name"])
            ws.cell(row=row, column=3, value=bug["test_area"])
            ws.cell(row=row, column=4, value=bug["severity"])
            ws.cell(row=row, column=5, value=bug["bug_type"])
            ws.cell(row=row, column=6, value=bug["expected"])
            
            # Truncate long actual text and clean it
            actual_text = str(bug["actual"])
            if len(actual_text) > 500:
                actual_text = actual_text[:500] + "..."
            # Remove problematic characters
            actual_text = actual_text.replace('\x00', '').replace('\n', ' ').replace('\r', ' ')
            ws.cell(row=row, column=7, value=actual_text)
            
            ws.cell(row=row, column=8, value=bug["environment"])
            ws.cell(row=row, column=9, value=bug["status"])
            ws.cell(row=row, column=10, value=bug["timestamp"])
            
            # Color code by severity
            severity_colors = {
                "High": "FFCCCC",
                "Medium": "FFFFCC", 
                "Low": "CCFFCC"
            }
            if bug["severity"] in severity_colors:
                fill = PatternFill(start_color=severity_colors[bug["severity"]], 
                                 end_color=severity_colors[bug["severity"]], fill_type="solid")
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        logger.info(f"Enhanced Excel bug report generated: {filename}")
        return filename
